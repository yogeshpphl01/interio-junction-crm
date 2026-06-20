"""
<module name="meta_import" layer="domain-pure">
  <purpose>
    Pure (no DB / no FastAPI / no auth) translation of a Meta Lead-Ads
    spreadsheet into CRM lead documents. Keeping this logic dependency-free makes
    it fully unit-testable (see tests/test_import_mapping.py) and reusable by the
    routers/imports.py HTTP endpoint.
  </purpose>
  <responsibilities>
    <item>Parse .xlsx / .csv bytes into normalized {header: value} rows.</item>
    <item>Map each Meta column to the corresponding CRM lead field.</item>
    <item>Humanize the free-text Meta form answers into a requirement brief.</item>
  </responsibilities>
</module>
"""
import io
import csv
import uuid
from datetime import datetime, timezone
from typing import Callable, Optional

import openpyxl


# <map>Meta 'number_of_bedrooms' answer -> CRM bhk_type.</map>
BHK_MAP = {"1": "1 BHK", "2": "2 BHK", "3": "3 BHK", "4": "4 BHK", "5": "5 BHK", "4+": "4 BHK", "5+": "5 BHK"}


def humanize(val) -> str:
    """'in_the_next_1_month_' -> 'In the next 1 month'."""
    if val is None:
        return ""
    s = str(val).strip().strip("_").replace("_", " ").strip()
    return (s[:1].upper() + s[1:]) if s else ""


def map_bedrooms(val) -> str:
    """Meta 'number_of_bedrooms' ('2', '4+', ...) -> CRM bhk_type."""
    if val is None:
        return "2 BHK"
    s = str(val).strip()
    if s in BHK_MAP:
        return BHK_MAP[s]
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits:
        n = int(digits[0])
        if 1 <= n <= 5:
            return f"{n} BHK"
        if n > 5:
            return "Villa"
    return "2 BHK"


def map_source(platform) -> str:
    """Meta 'platform' (ig/fb) -> CRM lead source label."""
    p = str(platform or "").strip().lower()
    if p in ("ig", "instagram"):
        return "Instagram"
    if p in ("fb", "facebook"):
        return "Facebook"
    return "Other"


def parse_created_time(val) -> Optional[str]:
    """Parse Meta's ISO timestamp (with +05:30 offset) into a UTC ISO string."""
    if val in (None, ""):
        return None
    if isinstance(val, datetime):
        dt = val
    else:
        s = str(val).strip().replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(s)
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat()


def clean_phone(val) -> str:
    """Trim whitespace inside a phone number ('+9180079 00780' -> '+918007900780')."""
    if val is None:
        return ""
    return "".join(str(val).split())


def to_bool(val) -> bool:
    return str(val).strip().lower() in ("true", "1", "yes", "y", "t")


def build_requirements(scope, timeline, priority, campaign) -> str:
    """Assemble a readable requirement brief from the Meta form answers."""
    parts = []
    if scope:
        parts.append(f"Scope: {humanize(scope)}.")
    if timeline:
        parts.append(f"Timeline: {humanize(timeline)}.")
    if priority:
        parts.append(f"Priority: {humanize(priority)}.")
    if campaign:
        parts.append(f"(Campaign: {campaign})")
    return " ".join(parts).strip()


def pick(row: dict, *exacts: str, contains: tuple = ()):
    """
    Resolve a value from a normalized row. Tries EXACT header names first (safe
    for short/ambiguous names like 'id' which is a substring of 'ad_id'), then
    optional substring matches.
    """
    for n in exacts:
        if n in row and row[n] not in (None, ""):
            return row[n]
    for sub in contains:
        for k, v in row.items():
            if sub in k and v not in (None, ""):
                return v
    return None


def _default_journey(ts: str) -> list[dict]:
    """Fallback stage-1 journey (router injects core.init_journey instead)."""
    return [{"stage": 1, "stage_name": "Captured", "entered_at": ts, "exited_at": None}]


def map_meta_row(
    row: dict,
    uploader_id: str,
    batch_id: str,
    fallback_ts: str,
    journey_factory: Callable[[str], list] = _default_journey,
) -> dict:
    """
    <function name="map_meta_row">
      Translate one normalized spreadsheet row into a full CRM lead document
      (stage 1, Active, Enquiry phase). `journey_factory` lets the router inject
      the canonical core.init_journey so the journey shape stays consistent.
    </function>
    """
    created = parse_created_time(pick(row, "created_time", contains=("created",))) or fallback_ts
    campaign = pick(row, "campaign_name", contains=("campaign_name",))
    scope = pick(row, "what's_the_scope_of_work", contains=("scope_of_work", "scope"))
    timeline = pick(row, "when_would_you_like_to_start_your_project?",
                    contains=("start_your_project", "when_would_you", "timeline"))
    priority = pick(row, "what_is_most_important_to_you?", contains=("most_important", "important_to_you"))
    platform = pick(row, "platform", contains=("platform",))
    organic = pick(row, "is_organic", contains=("organic",))
    meta_id = pick(row, "id", "lead_id")
    return {
        "id": str(uuid.uuid4()),
        "full_name": (str(pick(row, "full_name", contains=("full_name",)) or "").strip() or "Unknown"),
        "email": (str(pick(row, "email", contains=("email",)) or "").strip() or None),
        "phone": clean_phone(pick(row, "whatsapp_number", "phone", "mobile_number", contains=("whatsapp", "mobile"))),
        "city": (str(pick(row, "city", contains=("city",)) or "").strip() or None),
        "address": None,
        "lead_type": "Retail Client",
        "source": map_source(platform),
        "bhk_type": map_bedrooms(pick(row, "number_of_bedrooms", contains=("bedroom",))),
        "kitchen_layout": "L-shape",
        "tentative_budget": 0,
        "requirements": build_requirements(scope, timeline, priority, campaign),
        "assigned_to": uploader_id,
        "created_by": uploader_id,
        "stage": 1,
        "status": "Active",
        "project_id": None,
        # --- journey (a freshly imported lead has only enquired) ---
        "lifecycle_phase": "Enquiry",
        "furthest_stage": 1,
        "journey": journey_factory(created),
        # --- Meta provenance ---
        "meta_lead_id": (str(meta_id).strip() if meta_id else None),
        "source_platform": (str(platform).strip() if platform else None),
        "source_campaign": (str(campaign).strip() if campaign else None),
        "scope_of_work": (str(scope).strip() if scope else None),
        "project_timeline": (str(timeline).strip() if timeline else None),
        "priority_pref": (str(priority).strip() if priority else None),
        "is_organic": to_bool(organic) if organic is not None else None,
        "import_batch_id": batch_id,
        "created_at": created,
        "updated_at": fallback_ts,
    }


def parse_spreadsheet(content: bytes, filename: str) -> list[dict]:
    """Return rows as {normalized_header: value} dicts, from .xlsx or .csv bytes."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or "").strip().lower(): v for k, v in r.items()} for r in reader]

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers: Optional[list[str]] = None
    out: list[dict] = []
    for r in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [(str(c).strip().lower() if c is not None else f"col{i}") for i, c in enumerate(r)]
            continue
        d = {h: (r[i] if i < len(r) else None) for i, h in enumerate(headers)}
        if all(v is None or str(v).strip() == "" for v in d.values()):
            continue  # skip fully-empty rows
        out.append(d)
    wb.close()
    return out
